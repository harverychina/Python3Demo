# -*- coding:utf8 -*-
# @Time：2021/11/17 2:23 下午
# @Author： Huang Jeff

import os
import datetime
from openpyxl import Workbook

file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "demo.xlsx")
wb = Workbook()
ws = wb.active

ws['A1'] = 42
ws.append([1, 2, 3])
ws['A3'] = datetime.datetime.now()

wb.save(file_path)



